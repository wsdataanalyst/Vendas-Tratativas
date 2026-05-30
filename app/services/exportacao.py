"""Geração de planilha Excel com todos os registros."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app import auth
from app.services import metricas_gestor as svc_mgestor
from app.services import tratativas as svc_tratativas
from app.services import vendas_performance as svc_perf
from app.services import vendas_tratativa as svc_vtrat
from app.utils.datetime_br import agora_iso


def _estilizar_cabecalho(ws, colunas: int):
    fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    for col in range(1, colunas + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _ajustar_larguras(ws, larguras: list[int]):
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def gerar_excel() -> tuple[BytesIO, str]:
    tratativas = svc_tratativas.listar()
    perf = svc_perf.listar()
    vtrat = svc_vtrat.listar()
    painel = svc_mgestor.painel_gestor()
    c = painel["consolidado"]
    m_perf = c["perf"]
    m_vtrat = c["vtrat"]
    m_trat = c["trat"]

    wb = Workbook()

    ws_perf = wb.active
    ws_perf.title = "Vendas Performance"
    cab_perf = [
        "Data",
        "Operador",
        "Pedido",
        "Cliente",
        "Vendedor",
        "Valor (R$)",
        "Status",
        "Previsão",
        "Concorrência",
        "Observação",
    ]
    ws_perf.append(cab_perf)
    _estilizar_cabecalho(ws_perf, len(cab_perf))
    for v in perf:
        ws_perf.append(
            [
                v.get("data_registro", ""),
                auth.nome_operador(v.get("registrado_por")),
                v.get("pedido", ""),
                v.get("cliente") or "",
                v.get("vendedor") or "",
                v.get("valor"),
                v.get("status_venda", ""),
                v.get("previsao_data") or "",
                v.get("concorrencia") or "",
                v.get("observacao") or "",
            ]
        )
    _ajustar_larguras(ws_perf, [18, 16, 14, 16, 14, 12, 18, 14, 22, 28])

    ws_vtrat = wb.create_sheet("Vendas Tratativa")
    cab_vt = [
        "Data",
        "ID Tratativa",
        "Orçamento tratativa",
        "Pedido",
        "Valor (R$)",
        "Resultado",
        "Setor",
        "Situação",
        "Observação",
    ]
    ws_vtrat.append(cab_vt)
    _estilizar_cabecalho(ws_vtrat, len(cab_vt))
    for v in vtrat:
        res = v.get("resultado_tratativa") or ""
        if res == "com_impacto":
            res = "Com impacto"
        elif res == "sem_impacto":
            res = "Sem impacto"
        ws_vtrat.append(
            [
                v.get("data_registro", ""),
                v.get("id_tratativa"),
                v.get("tratativa_orcamento") or "",
                v.get("pedido", ""),
                v.get("valor"),
                res,
                v.get("tratativa_setor") or "",
                v.get("tratativa_situacao") or "",
                v.get("observacao") or "",
            ]
        )
    _ajustar_larguras(ws_vtrat, [18, 10, 14, 14, 12, 14, 14, 32, 28])

    ws_trat = wb.create_sheet("Tratativas")
    cab_trat = [
        "Data",
        "ID",
        "Operador",
        "Nº Orçamento",
        "Setor",
        "Situação",
        "Código item",
        "Tempo solução",
        "R$ Impacto",
        "Status",
        "Observação",
    ]
    ws_trat.append(cab_trat)
    _estilizar_cabecalho(ws_trat, len(cab_trat))
    for t in tratativas:
        ws_trat.append(
            [
                t.get("data_registro", ""),
                t.get("id"),
                auth.nome_operador(t.get("registrado_por")),
                t.get("numero_orcamento") or "",
                t.get("setor", ""),
                t.get("situacao", ""),
                t.get("codigo_item") or "",
                t.get("tempo_solucao") or "",
                t.get("impacto_reais"),
                t.get("status", ""),
                t.get("observacao") or "",
            ]
        )
    _ajustar_larguras(ws_trat, [18, 6, 16, 14, 14, 38, 12, 14, 12, 22, 30])

    ws_resumo = wb.create_sheet("Resumo", 0)
    ws_resumo.append(["Vendas & Tratativas — Exportação"])
    ws_resumo["A1"].font = Font(bold=True, size=14)
    ws_resumo.append(["Exportado em (Brasília)", agora_iso()])
    ws_resumo.append([])
    ws_resumo.append(["Vendas Performance (direto)"])
    ws_resumo.append(["Total", m_perf["total"]])
    ws_resumo.append(["Realizadas", m_perf["realizadas"]])
    ws_resumo.append(["Em andamento", m_perf["em_andamento"]])
    ws_resumo.append(["Perdidos", m_perf["perdidos"]])
    ws_resumo.append(["Total vendas R$ (Performance + Tratativa)", m_perf["valor_vendas_todas"]])
    ws_resumo.append([])
    for op_id, dados in painel["por_operador"].items():
        ws_resumo.append([dados["nome"]])
        ws_resumo.append(["  Performance realizadas", dados["perf"]["realizadas"]])
        ws_resumo.append(["  Via tratativa", dados["vtrat"]["total"]])
        ws_resumo.append(["  Tratativas abertas", dados["trat"]["em_andamento"]])
        ws_resumo.append([])
    ws_resumo.append(["Vendas via tratativa"])
    ws_resumo.append(["Total convertidas", m_vtrat["total"]])
    ws_resumo.append(["Sem impacto", m_vtrat["sem_impacto"]])
    ws_resumo.append(["Com impacto", m_vtrat["com_impacto"]])
    ws_resumo.append([])
    ws_resumo.append(["Tratativas"])
    ws_resumo.append(["Total", m_trat["total"]])
    ws_resumo.append(["Em aberto", m_trat["em_andamento"]])
    ws_resumo.append(["Resolvidas", m_trat["resolvidas"]])
    _ajustar_larguras(ws_resumo, [32, 22])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nome = f"vendas_tratativas_{agora_iso().replace(':', '').replace(' ', '_')}.xlsx"
    return buffer, nome
